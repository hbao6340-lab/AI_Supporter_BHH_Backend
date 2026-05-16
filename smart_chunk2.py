"""Parse Excel with proper openpyxl row structure."""
from pathlib import Path
import sys

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, 'backend')

import openpyxl
import re

DATA = Path('backend/knowledge/data')
XLSX = DATA / 'DANH SÁCH TỔNG THỂ (HC).xlsx'

wb = openpyxl.load_workbook(str(XLSX), data_only=True)
print("Sheets:", wb.sheetnames)

ws = wb['LĨNH VỰC']

SECTION_RE = re.compile(r'^[IVX]+[\.\s]', re.I)  # I. III. IV. V.

sections = []
current_section_rows = []

for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row_idx > 100:
        break
    
    vals = [str(v).strip() if v is not None else '' for v in row]
    first_cell = vals[0] if vals else ''
    second_cell = vals[1] if len(vals) > 1 else ''
    
    # Empty row? skip
    if not any(v.strip() for v in vals[:6]):
        continue
    
    # Column header row?
    combined = ' '.join(vals[:6]).lower()
    is_header = 'họ và tên' in combined or 'stt' in combined
    
    # Section marker?
    if SECTION_RE.match(first_cell) and not is_header:
        # Save previous section
        if current_section_rows:
            sections.append(current_section_rows)
        title = ' | '.join(v for v in vals[:4] if v.strip())
        current_section_rows = [f"[PHẦN] {first_cell}. {second_cell or title}"]
    elif current_section_rows and not is_header and vals[0]:
        row_text = ' | '.join(v for v in vals[:9] if v.strip())
        if row_text:
            current_section_rows.append(row_text)

# Save last section
if current_section_rows:
    sections.append(current_section_rows)

print(f"\nSections found: {len(sections)}")
for i, sec_rows in enumerate(sections):
    sec_title = sec_rows[0][:80] if sec_rows else '?'
    total_chars = sum(len(r) for r in sec_rows)
    print(f"\n[{i+1}] {sec_title}")
    print(f"     rows={len(sec_rows)}, chars={total_chars}")
    if len(sec_rows) <= 5:
        for r in sec_rows:
            print(f"   > {r[:120]}")
    else:
        print(f"   First: {sec_rows[1][:120] if len(sec_rows)>1 else '(data)'}")
        print(f"   Last:  {sec_rows[-1][:80]}")
