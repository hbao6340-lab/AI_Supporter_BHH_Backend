"""
Smart Excel chunking by section using openpyxl to read structure.
This properly preserves section headers + their member records as single chunks.
"""
from pathlib import Path
import sys
import re
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, 'backend')

import openpyxl
from knowledge.retriever import KnowledgeRetriever
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

DATA = Path('backend/knowledge/data')
XLSX_PATH = DATA / 'DANH SÁCH TỔNG THỂ (HC).xlsx'
print(f"Loading: {XLSX_PATH}")

wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
print(f"Sheets: {wb.sheetnames}")

# ─── READ EACH SHEET AS SECTIONS ──────────────────────────────────────────────
def parse_sheet(ws, sheet_name):
    """Parse a sheet, grouping rows by their identifiable section headers."""
    # Detect section headers: cell values like I., II., III., IV., V., etc.
    # with UB/MTTQ/P related content
    SECTION_MARKER_RE = re.compile(r'^[IVX]+[\s\.]', re.IGNORECASE)
    
    current_section = None
    current_section_title = None
    section_rows = defaultdict(list)
    
    # Track full text for each section
    section_texts = {}
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Skip completely empty rows
        if not any(v is not None and str(v).strip() for v in row):
            continue
        
        # Skip column header rows (first row in sheet)
        row_str = ' | '.join(str(v) if v is not None else '' for v in row)
        row_lower = row_str.lower()
        
        # Detect if this is a column header row
        if 'stt' in row_lower[:100] or 'họ và tên' in row_lower or 'họ tên' in row_lower:
            # This is a column header, not a data row - skip or add to current
            if current_section:
                section_rows[current_section].append(row_str)
            continue
        
        # Check for section header marker (I., II. III. etc.)
        first_cell = str(row[0]).strip() if row[0] is not None else ''
        is_section = bool(SECTION_MARKER_RE.match(first_cell))
        
        if is_section:
            # This is a section header row
            # Create new section based on the header content
            section_title = ' | '.join(str(v) if v is not None else '' for v in row if v).strip()
            
            # Check if this is a meaningful organizational section
            # by looking for keywords in the row content
            section_key = f"S{row_idx}"
            current_section = section_key
            current_section_title = section_title
            
            # Add section header text
            section_rows[section_key].append(f"[PHẦN] {section_title}")
        elif current_section:
            # This is a data row in the current section
            section_rows[current_section].append(row_str)
    
    return section_rows, section_texts


# Parse first sheet
for sheet_name in wb.sheetnames[:1]:  # Just first sheet
    ws = wb[sheet_name]
    section_rows, section_texts = parse_sheet(ws, sheet_name)
    print(f"\nSheet '{sheet_name}': {len(section_rows)} sections")
    
    for sec_key, rows in section_rows.items():
        combined = '\n'.join(rows)
        # Find section title from first row if it starts with [PHẦN]
        title = rows[0].replace('[PHẦN] ', '') if rows and rows[0].startswith('[PHẦN]') else sec_key
        print(f"\n  [{sec_key}] {title[:80] if len(title) > 80 else title}")
        print(f"    Rows: {len(rows)}, Chars: {len(combined)}")
        print(f"    Sample: {combined[:200]}")
