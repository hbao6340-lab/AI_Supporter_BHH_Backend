"""Inspect raw openpyxl rows."""
from pathlib import Path
import sys

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, 'backend')

import openpyxl

DATA = Path('backend/knowledge/data')
XLSX = DATA / 'DANH SÁCH TỔNG THỂ (HC).xlsx'

wb = openpyxl.load_workbook(str(XLSX), data_only=True)
print("Raw sheetnames:", [repr(s) for s in wb.sheetnames])

# Try first sheet name
ws = wb[wb.sheetnames[0]]
print(f"\nSheet: {repr(ws.title)}")

print("\nFirst 60 rows (first 4 values each):")
for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row_idx > 60:
        break
    vals = []
    for v in row[:6]:
        if v is not None:
            vals.append(str(v).strip()[:30])
        else:
            vals.append('')
    if any(v for v in vals):
        print(f"  R{row_idx:3d}: {vals}")
